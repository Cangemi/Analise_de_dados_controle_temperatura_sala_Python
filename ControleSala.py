import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series

dft = pd.DataFrame(columns=['Lim Max %RH','Lim Min %RH','Lim Max Temp','Lim Min Temp','Data','Min %RH', 'Max %RH', 'Min Temp', 'Max Temp', 'Media %RH', 'Media Temp'])


def criar_grafico(workbook, sheet_name, min_col_x,min_col_y,cell,min,max):
    # Seleciona a aba desejada
    sheet = workbook[sheet_name]
    
    
    # Define os dados para o gráfico
    x_data = Reference(sheet, min_col=min_col_x, min_row=2, max_row=sheet.max_row)
    y_data = Reference(sheet, min_col=min_col_y, min_row=2, max_row=sheet.max_row)
    y_min = Reference(sheet, min_col=min, min_row=2, max_row=sheet.max_row)
    y_max = Reference(sheet, min_col=max, min_row=2, max_row=sheet.max_row)

    # Cria um objeto de gráfico de linha
    chart = LineChart()
    chart.title = sheet.cell(row=1, column=min_col_y).value
    chart.x_axis.title = "Tempo"
    chart.y_axis.title = sheet.cell(row=1, column=min_col_y).value
    
    # Adiciona os dados ao gráfico
    chart.add_data(y_data, titles_from_data=True)
    chart.add_data(y_max, titles_from_data=True)
    chart.add_data(y_min, titles_from_data=True)
    chart.set_categories(x_data)
    
    # Ajusta as cores das linhas
    for idx, series in enumerate(chart.series):
        if idx == 0:  # Primeira série (y_data)
            series.graphicalProperties.line.solidFill = "000000"  # Cor preta
        elif idx == 1:  # Segunda série (y_max)
            series.graphicalProperties.line.solidFill = "FF0000"  # Cor vermelha
        elif idx == 2:  # Terceira série (y_min)
            series.graphicalProperties.line.solidFill = "FF0000"  # Cor vermelha
    
    # Define o tamanho do gráfico
    chart.height = 10
    chart.width = 15
    
    # Define o intervalo de escala do eixo Y
    valor_min = int(sheet.cell(row=2, column=min).value)
    valor_max = int(sheet.cell(row=2, column=max).value)
    chart.y_axis.scaling.min = valor_min - 2  # Mínimo do eixo Y
    chart.y_axis.scaling.max = valor_max + 2  # Máximo do eixo Y
    
    # Adiciona o gráfico à aba
    sheet.add_chart(chart, cell)

def salvar(workbook, df, nome_aba):
    # Adiciona uma nova aba com o nome fornecido
    sheet = workbook.create_sheet(title=nome_aba)
    
    # Converte o DataFrame em uma lista de listas de valores
    data = [df.columns.tolist()] + df.values.tolist()
    
    # Adiciona os dados à aba
    for row in data:
        sheet.append(row)

def processamento(caminho_arquivo, workbook):
    try:
        #acessa variavel global
        global dft
        
        # Lê o arquivo .txt como um DataFrame pandas
        df = pd.read_csv(caminho_arquivo, delimiter='\t')
        
        # Pega o nome da aba a partir do nome do arquivo
        nome_aba = os.path.basename(caminho_arquivo)
        nome_aba = os.path.splitext(nome_aba)[0]
        
        df['%RH'] = df['%RH'].str.replace('%RH', '').str.replace(',', '.').astype(float)
        df['Temp'] = df['Temp'].str.replace(',', '.').astype(float)
        
        df.insert(1, 'Lim Max %RH', 70)
        df.insert(2, 'Lim Min %RH', 40)
        df.insert(3, 'Lim Max Temp', 26)
        df.insert(4, 'Lim Min Temp', 20)
        
        MinRH     = df['%RH'].min()
        MaxRH     = df['%RH'].max()
        MinTemp   = df['Temp'].min()
        MaxTemp   = df['Temp'].max()
        MediaRH   = df['%RH'].mean()
        MediaTemp = df['Temp'].mean()
        
        df.loc[0, 'Min %RH']    = MinRH
        df.loc[0, 'Max %RH']    = MaxRH
        df.loc[0, 'Min Temp']   = MinTemp
        df.loc[0, 'Max Temp']   = MaxTemp
        df.loc[0, 'Media %RH']  = MediaRH
        df.loc[0, 'Media Temp'] = MediaTemp
        
        
        dft.loc[len(dft)] = [70,40,26,20,nome_aba,MinRH, MaxRH, MinTemp, MaxTemp, MediaRH, MediaTemp]
        
        
        # Salva o DataFrame em uma aba com o nome obtido
        salvar(workbook, df, nome_aba)
        
        criar_grafico(workbook, nome_aba,6,8,"J4",5,4)
        criar_grafico(workbook, nome_aba,6,7,"J24",3,2)
            
        df = None
    
    except Exception as e:
        print(f"Erro ao processar o arquivo: {caminho_arquivo}. Detalhes do erro: {str(e)}")


def abrir_planilhas_diretorio(diretorio):

    # Cria uma nova instância da classe Workbook
    workbook = Workbook()
    # Remove a planilha padrão
    workbook.remove(workbook.active)
    
    
    # Verifica se o diretório existe
    if os.path.exists(diretorio):
        
        # Define o caminho de saída para o arquivo Excel
        caminho_saida = os.path.join(diretorio, 'consolidado_mes.xlsx')
        #Verifica se existe o arquivo
        if os.path.exists(caminho_saida):
            os.remove(caminho_saida)
        
        # Percorre todos os arquivos no diretório
        for arquivo in os.listdir(diretorio):
            #adiciona o caminho completo do arquivo à lista
            caminho_arquivo = os.path.join(diretorio, arquivo)
            # Processa o arquivo
            processamento(caminho_arquivo,workbook)
            
        
        MinRH     = dft['Min %RH'].min()
        MaxRH     = dft['Max %RH'].max()
        MinTemp   = dft['Min Temp'].min()
        MaxTemp   = dft['Max Temp'].max()
        MediaRH   = dft['Media %RH'].mean()
        MediaTemp = dft['Media Temp'].mean()
        
        dft['Lim Max %RH']    = 70
        dft['Lim Min %RH']    = 40
        dft['Lim Max Temp']    = 26
        dft['Lim Min Temp']    = 20
        dft.loc[0, 'Min Mes %RH']    = MinRH
        dft.loc[0, 'Max Mes %RH']    = MaxRH
        dft.loc[0, 'Min Mes Temp']   = MinTemp
        dft.loc[0, 'Max Mes Temp']   = MaxTemp
        dft.loc[0, 'Media Mes %RH']  = MediaRH
        dft.loc[0, 'Media Mes Temp'] = MediaTemp
        
        # Salvar aba geral
        salvar(workbook, dft, 'Geral')
        
        # Obter a aba "geral"
        aba_geral = workbook["Geral"]

        # Remover a aba "geral" do arquivo
        workbook.remove(aba_geral)

        # Inserir a aba "geral" como a primeira aba
        workbook._sheets.insert(0, aba_geral)
        
        
        criar_grafico(workbook, 'Geral',5,11,"M4",4,3)
        criar_grafico(workbook, 'Geral',5,10,"M24",2,1)
        
        # Salva o workbook como um arquivo Excel                 
        workbook.save(caminho_saida)
        print(f"Planilha consolidada '{caminho_saida}' criada com sucesso.")
    
    else:
        print("O diretório especificado não existe.")
        
     

if __name__ == "__main__":
    diretorio = input("Digite o diretório: ")
    abrir_planilhas_diretorio(diretorio)
